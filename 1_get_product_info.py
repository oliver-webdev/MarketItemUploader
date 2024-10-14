from get_token import get_token
from read_source_exl import read_excel_data
from get_amazon_product_data import get_product_detail
from upload_image import upload_product_image
import openpyxl
import os

def main():
    # 토큰 생성
    token = get_token(client_id='wlzNpmIHUiTBUzf6Bor7C', client_secret='$2a$04$IUvzowqvAMl/LbLCGNUgke')
    print(f"생성된 토큰: {token}")
    
    # 엑셀 파일 경로
    file_path = 'product_upload_source.xlsx'        # 아마존 상품 URL 목록 파일 
    template_path = 'product_upload_template.xlsx'  # 템플릿 파일
    output_path = 'product_upload_data.xlsx'        # 마켓 API로 업로드할 파일
    
    data_list = read_excel_data(file_path)
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # 데이터는 2행부터 쓰기 시작합니다
    row_index = 2
    
    # 1. 엑셀파일 읽기
    for data in data_list:
        amazon_url = data['amazon_url']
        
        try:
            # 아마존 상품 정보 가져오기
            res = get_product_detail(amazon_url)
        except Exception as e:
            print(f"정보 가져오기 중 에러 발생: {e}")

        try:
            # 상품 이미지 네이버 업로드
            arr_img_path = upload_product_image(res.get('img_path'), token)
        except Exception as e:
            print(f"이미지 업로드 중 에러 발생: {e}")

        amazon_title = res.get('title')
        amazon_desc = res.get('desc')
        amazon_table = res.get('html_table')
        amazon_price = res.get('price')

        # 각 데이터를 지정된 셀에 씁니다.
        ws[f'A{row_index}'] = amazon_title
        ws[f'B{row_index}'] = amazon_price
        ws[f'C{row_index}'] = amazon_desc
        ws[f'D{row_index}'] = amazon_table
        ws[f'E{row_index}'] = arr_img_path

        row_index += 1  # 다음 행으로 이동

    # 기존 존재하던 파일 삭제
    if os.path.exists(output_path):
        os.remove(output_path)
        print(f"{output_path} 파일 삭제")

    # 파일 쓰기
    wb.save(output_path)
    print(f"{output_path} 파일에 저장되었습니다.")

if __name__ == "__main__":
    main()
